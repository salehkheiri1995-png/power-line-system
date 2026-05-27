import openpyxl
from typing import List, Dict, Any
import re

def parse_excel(file_path: str) -> List[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    records = []
    for sheet_name in workbook.sheetnames:
        # تعیین نوع برنامه از نام شیت (فرض می‌کنیم حاوی "سرد" یا "گرم" است)
        program_type = "سرد" if "سرد" in sheet_name else "گرم"

        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 4:
            continue  # نیاز به حداقل 4 ردیف (سه سطر اول معمولاً عنوان‌ها)

        # فرض می‌کنیم سطر چهارم (ایندکس 3) هدر است
        headers = rows[3]
        if not headers:
            continue
        # داده‌ها از سطر پنجم به بعد (ایندکس 4)
        for row in rows[4:]:
            if not row or all(cell is None for cell in row):
                continue
            record = {
                "program_type": program_type,
                "code": None,
                "line_name": None,
                "voltage_level": None,
                "work_description": None,
                "tower_number": None,
                "location": None,
                "pm_date": None,
                "execution_date": None,
                "team_count": None,
                "personnel_count": None,
                "supervisor": None,
                "quantity": None,
                "unit": None,
                "tower_number2": None,
                "title_of_work": None,
                "extra_tower_number": None,
            }
            for idx, header in enumerate(headers):
                if header is None:
                    continue
                val = row[idx] if idx < len(row) else None
                # تبدیل به رشته و حذف فضای خالی
                if val is not None:
                    val = str(val).strip()
                # نگاشت هدر به فیلدهای مدل
                h = str(header).strip()
                if "کد" in h and "نام" not in h:
                    record["code"] = val
                elif "نام خط" in h:
                    record["line_name"] = val
                elif "سطح ولتاژ" in h:
                    record["voltage_level"] = val
                elif "نوع برنامه" in h:
                    record["program_type"] = val if val else program_type
                elif "شرح انجام کار" in h:
                    record["work_description"] = val
                elif "عنوان کار" in h and program_type == "گرم":
                    record["title_of_work"] = val
                elif "شماره دکل" in h:
                    # ممکن است دو ستون داشته باشیم
                    if record["tower_number"] is None:
                        record["tower_number"] = val
                    else:
                        record["tower_number2"] = val
                elif "موقعیت" in h:
                    record["location"] = val
                elif "تاریخ pm" in h:
                    record["pm_date"] = val
                elif "تاریخ انجام" in h:
                    record["execution_date"] = val
                elif "تعداد اکیپ" in h:
                    try: record["team_count"] = float(val) if val else None
                    except: pass
                elif "تعداد نفرات" in h:
                    try: record["personnel_count"] = float(val) if val else None
                    except: pass
                elif "نام سرپرست" in h:
                    record["supervisor"] = val
                elif "مقدار" in h:
                    try: record["quantity"] = float(val) if val else None
                    except: pass
                elif "واحد" in h:
                    record["unit"] = val
                # اگر ستون اضافی شماره دکل در گرم باشد
                if program_type == "گرم" and "شماره" in h and record["tower_number"] and record["tower_number"] != val:
                    record["extra_tower_number"] = val
            records.append(record)
    return records
